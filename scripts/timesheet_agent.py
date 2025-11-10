import argparse  # Handles command-line interface parsing
import re  # Provides regular expressions for pattern matching
import shutil  # Offers high-level file operations (copy)
from datetime import date, datetime  # Supplies date/time classes for parsing cells
from pathlib import Path  # Enables filesystem path operations in an OS-agnostic way
from typing import List  # Provides type hints for better readability

import pandas as pd  # Core data manipulation library
from openpyxl import load_workbook  # Used to read/write Excel workbooks
from openpyxl.styles import PatternFill  # Controls cell background fills

EXPECTED_SOURCE_COLUMNS = [  # Enforces the exact order of expected columns
    "ActivityDescription",  # Raw student/service description field
    "TherapistName",  # Therapist name column
    "Date",  # Date column in source file (later renamed)
    "PlacementDesc",  # Service/placement description
    "RateName",  # Rate label (unused but preserved)
    "EntryQuantity",  # Logged hours
    "ChargeRate",  # Hourly rate column
    "TotalCharge",  # Total charge column
]  # End list of source columns

TARGET_COLUMNS = [  # Defines output order for the invoice sheet
    "Therapist Name",  # Therapist name target column
    "Date of Service",  # Date column formatted as mm/dd/yyyy
    "PA Cyber Student ID",  # Student ID extracted from ActivityDescription
    "Student Name",  # Student name derived from ActivityDescription
    "Service",  # Placement description column
    "Therapy ONLY Session Minutes on IEP ",  # Minutes column required by template
]  # End target column list

STUDENT_ID_PATTERN = re.compile(r"^[A-Za-z]{2}\d{6}$")  # Matches two letters plus six digits (e.g., AB123456)


def load_source_dataframe(path: Path) -> pd.DataFrame:
    """Load the TimesheetPortal workbook and normalize it into DF_source."""

    raw = pd.read_excel(path, header=None)  # Read the sheet without predefined headers
    if raw.shape[0] < 3:  # Verify there are enough rows to contain data
        raise ValueError("Source workbook must contain at least three rows (title, headers, data).")  # Fail fast

    headers = raw.iloc[1].tolist()  # Second row contains the header labels
    df_source = raw.iloc[2:].copy()  # Actual data starts on the third row
    df_source.columns = headers  # Apply column names
    df_source = df_source.dropna(how="all").reset_index(drop=True)  # Remove fully empty rows

    missing = [col for col in EXPECTED_SOURCE_COLUMNS if col not in df_source.columns]  # Check for required columns
    if missing:  # If any columns are missing
        raise ValueError(f"Missing expected columns in source file: {missing}")  # Inform caller

    df_source = df_source[EXPECTED_SOURCE_COLUMNS].copy()  # Reorder columns consistently
    return df_source  # Return normalized dataframe


def dedupe_tokens(tokens: List[str]) -> List[str]:
    """Remove duplicate tokens while preserving original order."""

    seen = set()  # Tracks tokens already emitted
    ordered = []  # Stores the tokens in output order
    for token in tokens:  # Loop through each token
        if token not in seen:  # Only process unseen tokens
            seen.add(token)  # Mark token as seen
            ordered.append(token)  # Append to ordered list
    return ordered  # Return unique tokens preserving order


def extract_student_fields(activity: str) -> tuple[str, str]:
    """Split ActivityDescription into (student_id, student_name)."""

    tokens = [token.strip() for token in str(activity).split() if token.strip()]  # Clean and tokenize the string
    student_id = ""  # Initialize default ID
    student_name = ""  # Initialize default name

    for index, token in enumerate(tokens):  # Inspect each token sequentially
        if STUDENT_ID_PATTERN.match(token):  # Look for the ID pattern
            student_id = token  # Store the ID when found
            left_tokens = tokens[:index]  # Tokens before the ID
            right_tokens = tokens[index + 1 :]  # Tokens after the ID

            if len(right_tokens) > len(left_tokens):  # Choose the longer side
                candidate = right_tokens  # Use right-side tokens
            elif len(right_tokens) == len(left_tokens):  # If equal length
                candidate = left_tokens  # Default to the left side
            else:
                candidate = left_tokens  # Left side is longer

            candidate = dedupe_tokens(candidate)  # Remove duplicate tokens
            student_name = " ".join(candidate).strip()  # Join into a single string
            break  # Exit loop after finding the ID

    return student_id, student_name  # Provide extracted ID/name pair


def build_target_dataframe(df_source: pd.DataFrame) -> pd.DataFrame:
    """Transform DF_source rows into the DF_target structure."""

    records = []  # Holds transformed dicts prior to DataFrame creation

    for _, row in df_source.iterrows():  # Iterate over every source row
        if row.isna().all():  # Skip rows that are entirely NaN
            continue  # Move to the next row

        student_id, student_name = extract_student_fields(row.get("ActivityDescription"))  # Derive student fields

        therapist_name = str(row.get("TherapistName") or "").strip()  # Clean therapist name
        date_value = row.get("Date")  # Fetch raw date
        date_str = ""  # Initialize formatted date string
        if pd.notna(date_value):  # Only format when date exists
            if isinstance(date_value, (pd.Timestamp, datetime, date)):  # Handle actual datetime objects
                date_str = pd.to_datetime(date_value).strftime("%m/%d/%Y")  # Format as mm/dd/yyyy
            else:
                date_str = str(date_value).strip()  # Fallback for string dates

        service_value = str(row.get("PlacementDesc") or "").strip()  # Normalize service text

        entry_quantity = row.get("EntryQuantity")  # Retrieve hours
        minutes_value = ""  # Initialize minutes cell
        if pd.notna(entry_quantity):  # Convert when hours exist
            minutes_value = round(float(entry_quantity) * 60, 2)  # Convert hours to minutes with rounding

        record = {  # Assemble a single output record
            "Therapist Name": therapist_name,  # Output therapist name
            "Date of Service": date_str,  # Output formatted date
            "PA Cyber Student ID": student_id,  # Output student ID
            "Student Name": student_name,  # Output student name
            "Service": service_value,  # Output service field
            "Therapy ONLY Session Minutes on IEP ": minutes_value,  # Output minutes value
        }
        records.append(record)  # Append record to master list

    df_target = pd.DataFrame(records, columns=TARGET_COLUMNS)  # Build DF_target using the defined column order
    return df_target  # Return the transformed dataframe


def write_to_template(df_target: pd.DataFrame, template_path: Path, output_path: Path) -> None:
    """Copy the template file and paste DF_target rows into the correct cells."""

    output_path.parent.mkdir(parents=True, exist_ok=True)  # Ensure the destination directory exists
    shutil.copy(template_path, output_path)  # Copy the template so we never edit the original

    workbook = load_workbook(output_path)  # Open the copied workbook
    worksheet = workbook.active  # Work with the active sheet in the template

    start_row = 7  # Data in the template begins at row 7
    column_letters = ["A", "B", "C", "D", "E", "F"]  # Target columns for the six DF_target fields

    for row_index, (_, row) in enumerate(df_target.iterrows(), start=start_row):  # Loop over each target row
        values = row.tolist()  # Convert row to a list for easy indexing
        for col_idx, value in enumerate(values):  # Loop through each column value
            cell = worksheet[f"{column_letters[col_idx]}{row_index}"]  # Compute the cell coordinate
            cell.value = value  # Write the value into the cell
            cell.fill = PatternFill(fill_type=None)  # Clear fills to avoid red background flags

    workbook.save(output_path)  # Persist all modifications to disk


def parse_args() -> argparse.Namespace:
    """Define and parse CLI arguments for the converter."""

    parser = argparse.ArgumentParser(description="Convert TimesheetPortal export into invoice format.")  # Create parser
    parser.add_argument("--source", required=True, help="Path to TimesheetPortal-Report-*.xlsx file.")  # Source path arg
    parser.add_argument("--template", required=True, help="Path to INV#*.xlsx template file.")  # Template path arg
    parser.add_argument(
        "--output",  # Name of the output flag
        required=True,  # Must be provided
        help="Path for the generated invoice workbook.",  # Help text for users
    )
    return parser.parse_args()  # Parse and return namespace


def main() -> None:
    """Entry point when executed as a script."""

    args = parse_args()  # Parse provided CLI arguments
    convert_timesheet(args.source, args.template, args.output)  # Run the conversion with provided paths


def convert_timesheet(source: str | Path, template: str | Path, output: str | Path) -> Path:
    """Orchestrate the full conversion from source file to invoice output."""

    source_path = Path(source).resolve()  # Resolve the source path to an absolute path
    template_path = Path(template).resolve()  # Resolve the template path
    output_path = Path(output).resolve()  # Resolve the output path

    df_source = load_source_dataframe(source_path)  # Load and normalize the source report
    df_target = build_target_dataframe(df_source)  # Transform source rows into template-ready rows
    write_to_template(df_target, template_path, output_path)  # Populate the copied template with data

    print(f"DF_source rows: {len(df_source)}")  # Inform user how many source rows were processed
    print(f"DF_target rows: {len(df_target)}")  # Inform user how many rows landed in the target
    print(f"Invoice saved to: {output_path}")  # Echo the output path for quick reference
    return output_path  # Return the resolved output path for callers


if __name__ == "__main__":  # Only execute when run directly
    main()  # Invoke CLI entry point
